from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from utils.helper import compare_experience

def postprocess_resume_report(report_path):
    """
    Post-process the Excel report at report_path:
    - Set header row background to green
    - Set extracted columns text color (green/red) based on match with expected
    - Auto-resize columns
    """
    wb = load_workbook(report_path)
    ws = wb.active

    # 1. Set header row background to green
    header_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    for cell in ws[1]:
        cell.fill = header_fill

    # 2. Set extracted columns text color (green/red)
    header = [cell.value for cell in ws[1]]
    extract_expected_pairs = [
        ('Expected Name', 'Extracted Name'),
        ('Expected Email', 'Extracted Email'),
        ('Expected Mobile', 'Extracted Mobile'),
        ('Expected Location', 'Extracted Location'),
        ('Expected Experience', 'Extracted Experience'),
        ('Expected Company', 'Extracted Company'),
    ]
    # For summary
    summary_data = []
    total_wrong_cells = 0
    total_extracted_cells = 0
    for exp_col, ext_col in extract_expected_pairs:
        if exp_col in header and ext_col in header:
            exp_idx = header.index(exp_col) + 1  # 1-based for openpyxl
            ext_idx = header.index(ext_col) + 1
            wrong_count = 0
            total_count = 0
            for row in range(2, ws.max_row + 1):
                exp_val = ws.cell(row=row, column=exp_idx).value
                ext_val = ws.cell(row=row, column=ext_idx).value
                total_count += 1
                # Use compare_experience for Experience, else strict equality
                if exp_col == 'Expected Experience':
                    is_match = compare_experience(exp_val, ext_val)
                else:
                    is_match = str(exp_val).strip().lower() == str(ext_val).strip().lower()
                if is_match:
                    ws.cell(row=row, column=ext_idx).font = Font(color='008000')  # Green
                else:
                    ws.cell(row=row, column=ext_idx).font = Font(color='FF0000')  # Red
                    wrong_count += 1
            percentage = (wrong_count / total_count * 100) if total_count else 0
            summary_data.append((
                ext_col.replace('Extracted ', ''),
                f"{wrong_count}/{total_count}",
                f"{percentage:.1f}%"
            ))
            total_wrong_cells += wrong_count
            total_extracted_cells += total_count

    # 3. Auto-resize columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 4. Add summary sheet
    summary_ws = wb.create_sheet('Summary', 0)
    # Beautification styles
    header_fill = PatternFill(start_color='B3ceFB', end_color='B3ceFB', fill_type='solid')  # Light blue
    cell_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    first_col_fill = PatternFill(start_color='F7B4AE', end_color='F7B4AE', fill_type='solid')  # Light red
    overall_fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')  # Light orange
    bold_font = Font(bold=True)
    
    # Write header
    summary_ws.append(['Fields Name', 'Wrong cells', 'Percentage'])
    for cell in summary_ws[1]:
        cell.fill = header_fill
        cell.font = bold_font
    
    # Write data
    for row_idx, (field, wrong_total, percent) in enumerate(summary_data, start=2):
        summary_ws.cell(row=row_idx, column=1, value=field)
        summary_ws.cell(row=row_idx, column=2, value=wrong_total)
        summary_ws.cell(row=row_idx, column=3, value=percent)
        for col_idx in range(1, 4):
            if col_idx == 1:
                summary_ws.cell(row=row_idx, column=col_idx).fill = first_col_fill
            else:
                summary_ws.cell(row=row_idx, column=col_idx).fill = cell_fill
            # Check if there are wrong cells (extract wrong count from "wrong/total" format)
            wrong_count = int(wrong_total.split('/')[0])
            if col_idx == 2 and wrong_count > 0:
                summary_ws.cell(row=row_idx, column=col_idx).font = Font(color='FF0000')
            else:
                summary_ws.cell(row=row_idx, column=col_idx)
    
    # Add overall statistics
    overall_wrong_percentage = (total_wrong_cells / total_extracted_cells * 100) if total_extracted_cells else 0
    
    # Add empty row
    summary_ws.append([])
    
    # Add overall statistics row
    summary_ws.append(['Overall Wrong Cells', f"{total_wrong_cells}/{total_extracted_cells}", f"{overall_wrong_percentage:.1f}%"])
    overall_row = len(summary_data) + 3  # +3 because of header + empty row + overall row
    for col_idx in range(1, 4):
        summary_ws.cell(row=overall_row, column=col_idx).fill = overall_fill
        summary_ws.cell(row=overall_row, column=col_idx).font = bold_font
        if col_idx == 2 and total_wrong_cells > 0:
            summary_ws.cell(row=overall_row, column=col_idx).font = Font(color='FF0000', bold=True)
    
    # Auto-resize summary columns
    for col in range(1, 4):
        max_length = 0
        for row in range(1, overall_row + 1):
            val = summary_ws.cell(row=row, column=col).value
            if val:
                max_length = max(max_length, len(str(val)))
        summary_ws.column_dimensions[get_column_letter(col)].width = max_length + 2

    wb.save(report_path) 